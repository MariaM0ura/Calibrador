
"""
pipeline.py — Módulo importável de calibragem de campanhas Amazon Ads.

Uso:
    from pipeline import rodar_calibragem

    resultado = rodar_calibragem(
        arquivo=io.BytesIO(bytes_do_xlsx),
        roas_target=4.0,
        budget_diario_sp=500.0,
        budget_diario_sb=0.0,
        budget_diario_sd=0.0,
        ...
    )

O parâmetro `arquivo` aceita:
  - str / Path  → caminho para o .xlsx no disco
  - bytes       → conteúdo bruto do arquivo
  - file-like   → BytesIO, UploadedFile do Streamlit, etc.

O parâmetro `on_progress` é opcional: fn(pct: float, msg: str) chamada
a cada etapa para atualizar barras de progresso externas.

Retorna dict com chaves:
  n_bids, n_budgets, n_placements, relatorio, workbook, abas_removidas
"""

import io
import warnings
from pathlib import Path

import openpyxl

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# ---------------------------------------------------------------------------
# Constantes de layout da planilha (1-indexed, padrão openpyxl)
# ---------------------------------------------------------------------------
_COL_ENTITY         = 2   # B
_COL_OPERATION      = 3   # C
_COL_CAMPAIGN_NAME  = 10  # J  (preenchido só em linhas Campaign)
_COL_CAMPAIGN_INFO  = 12  # L  (informational — preenchido em todas as linhas)
_COL_BUDGET         = 21  # U
_COL_BID            = 28  # AB
_COL_PLACEMENT_TYPE = 34  # AH (ex: "Placement Top")
_COL_PLACEMENT_PCT  = 35  # AI
_COL_CLICKS         = 43  # AQ
_COL_SALES          = 46  # AT
_COL_ROAS           = 52  # AZ

_ABA_SP   = "Sponsored Products Campaigns"
_ABAS_RAS = ("RAS Campaigns", "RAS Search Term Report")
_PLACEMENT_MAXIMO = 900.0

# Por produto: nomes de aba, colunas de budget (índices estáveis no Bulk Sheet),
# entidades de linha com bid editável, e fallback de colunas bid/cliques/roas
# caso a linha 1 não traga cabeçalhos reconhecíveis.
_SHEET_CALIBRATION_CONFIG = {
    "sp": {
        "sheet_name": "Sponsored Products Campaigns",
        "campaign_col": 10,
        "budget_col": 21,
        "sales_col": 46,
        "roas_col": 52,
        "bid_entities": ("Keyword", "Product Targeting"),
        "bid_col_fb": 28,
        "clicks_col_fb": 43,
        "roas_col_fb": 52,
    },
    "sb": {
        "sheet_name": "Sponsored Brands Campaigns",
        "campaign_col": 10,
        "budget_col": 19,
        "sales_col": 45,
        "roas_col": 51,
        "bid_entities": ("Keyword", "Brand Keyword", "Product Targeting"),
        "bid_col_fb": 27,
        "clicks_col_fb": 42,
        "roas_col_fb": 51,
    },
    "sd": {
        "sheet_name": "Sponsored Display Campaigns",
        "campaign_col": 9,
        "budget_col": 21,
        "sales_col": 35,
        "roas_col": 41,
        "bid_entities": (
            "Keyword",
            "Product Targeting",
            "Audience Targeting",
            "Contextual Targeting",
            "Views Remarketing",
            "Purchases Remarketing",
            "Amazon Audiences",
        ),
        "bid_col_fb": 26,
        "clicks_col_fb": 32,
        "roas_col_fb": 41,
    },
}


# ---------------------------------------------------------------------------
# Helpers internos
# ---------------------------------------------------------------------------

def _to_float(val, default=0.0):
    if val is None or val == "":
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def _campaign_name(ws, row):
    name = ws.cell(row, _COL_CAMPAIGN_NAME).value or ws.cell(row, _COL_CAMPAIGN_INFO).value
    return (name or "").strip()


def _header_scan(ws, max_cols=220):
    out = []
    limit = max(max_cols, (ws.max_column or 0) + 1)
    for c in range(1, limit + 1):
        raw = ws.cell(1, c).value
        h = (str(raw).strip().lower() if raw is not None else "")
        out.append((c, h, str(raw).strip() if raw is not None else ""))
    return out


def _resolve_bulk_columns(ws, fallback_cfg=None):
    """
    Resolve colunas a partir da linha 1 do Bulk Sheet (inglês).
    fallback_cfg: dict opcional com bid_col_fb, clicks_col_fb, roas_col_fb, entity (default 2).
    """
    headers = _header_scan(ws)
    cmap = {}

    for c, h, _ in headers:
        if h == "entity":
            cmap["entity"] = c
            break
    for c, h, _ in headers:
        if h == "operation":
            cmap["operation"] = c
            break

    bid_c = None
    for c, h, _ in headers:
        if h == "max bid" or h.endswith(" max bid"):
            bid_c = c
            break
    if bid_c is None:
        for c, h, _ in headers:
            if h == "bid":
                bid_c = c
                break
    if bid_c is not None:
        cmap["bid"] = bid_c

    for c, h, _ in headers:
        if h == "clicks":
            cmap["clicks"] = c
            break
    if "clicks" not in cmap:
        for c, h, _ in headers:
            if "clicks" in h and "cost per" not in h and "click-through" not in h:
                cmap["clicks"] = c
                break

    roas_cands = [(c, h) for c, h, _ in headers if "roas" in h]
    if roas_cands:
        pref = [c for c, h in roas_cands if "14" in h and "sales" in h]
        cmap["roas"] = pref[0] if pref else roas_cands[0][0]

    for c, h, _ in headers:
        if "campaign name" in h and "informational" not in h:
            cmap["campaign_name"] = c
            break
    for c, h, _ in headers:
        if "informational" in h and "campaign name" in h:
            cmap["campaign_name_informational"] = c
            break

    fb = fallback_cfg or {}
    cmap.setdefault("entity", fb.get("entity_fb", _COL_ENTITY))
    cmap.setdefault("operation", fb.get("operation_fb", _COL_OPERATION))
    cmap.setdefault("bid", fb.get("bid_col_fb", _COL_BID))
    cmap.setdefault("clicks", fb.get("clicks_col_fb", _COL_CLICKS))
    cmap.setdefault("roas", fb.get("roas_col_fb", _COL_ROAS))

    return cmap


def _campaign_display_name(ws, row, cmap, fallback_campaign_col):
    name = None
    cn = cmap.get("campaign_name")
    if cn:
        name = ws.cell(row, cn).value
    if not name and fallback_campaign_col:
        name = ws.cell(row, fallback_campaign_col).value
    ci = cmap.get("campaign_name_informational")
    if not name and ci:
        name = ws.cell(row, ci).value
    return (name or "").strip()


def calcular_ajuste_roas(roas: float, target: float):
    """
    Retorna (fator, motivo) para um dado ROAS e target.

    fator > 1.0 → aumento  |  fator < 1.0 → redução  |  1.0 → sem ajuste
    Exportado para testes ou uso externo.
    """
    if target <= 0 or roas < 0:
        return 1.0, "Dados inválidos (target ≤ 0 ou ROAS negativo)"

    desvio = (roas - target) / target

    if roas > target:
        if desvio < 0.15:
            return 1.05, f"ROAS {roas:.2f} acima do target em {desvio*100:.1f}% → +5%"
        elif desvio < 0.25:
            return 1.07, f"ROAS {roas:.2f} acima do target em {desvio*100:.1f}% → +7%"
        elif desvio < 0.50:
            return 1.10, f"ROAS {roas:.2f} acima do target em {desvio*100:.1f}% → +10%"
        elif desvio < 1.00:
            return 1.15, f"ROAS {roas:.2f} acima do target em {desvio*100:.1f}% → +15%"
        else:
            return 1.20, f"ROAS {roas:.2f} acima do target em {desvio*100:.1f}% → +20%"

    elif roas < target:
        if desvio > -0.15:
            return 0.90, f"ROAS {roas:.2f} abaixo do target em {abs(desvio)*100:.1f}% → -10%"
        elif desvio > -0.25:
            return 0.85, f"ROAS {roas:.2f} abaixo do target em {abs(desvio)*100:.1f}% → -15%"
        elif desvio > -0.50:
            return 0.80, f"ROAS {roas:.2f} abaixo do target em {abs(desvio)*100:.1f}% → -20%"
        else:
            return 0.80, f"ROAS {roas:.2f} muito abaixo do target em {abs(desvio)*100:.1f}% → -20%"

    return 1.0, f"ROAS {roas:.2f} igual ao target → sem ajuste"


# ---------------------------------------------------------------------------
# Módulo 1 — Bid (por aba SP / SB / SD)
# ---------------------------------------------------------------------------

def _modulo_bid_sheet(ws, sheet_name, relatorio, cfg, sheet_cfg):
    cmap = _resolve_bulk_columns(ws, sheet_cfg)
    ec = cmap["entity"]
    bid_c = cmap["bid"]
    clk_c = cmap["clicks"]
    roa_c = cmap["roas"]
    op_c = cmap["operation"]
    bid_entities = sheet_cfg["bid_entities"]

    ajustados = 0
    roas_target = cfg["roas_target"]
    bid_maximo = cfg["bid_maximo"]
    dias = cfg["dias"]

    for row in range(2, ws.max_row + 1):
        entity = ws.cell(row, ec).value
        if entity not in bid_entities:
            continue

        bid_atual = _to_float(ws.cell(row, bid_c).value)
        if bid_atual <= 0:
            continue

        clicks = _to_float(ws.cell(row, clk_c).value)
        roas = _to_float(ws.cell(row, roa_c).value)
        camp = _campaign_display_name(ws, row, cmap, sheet_cfg["campaign_col"])

        if clicks < (5 * dias):
            novo_bid = bid_atual * 1.05
            motivo = f"Baixo volume de clicks ({int(clicks)} cliques < threshold {5 * dias})"
        else:
            fator, motivo = calcular_ajuste_roas(roas, roas_target)
            novo_bid = bid_atual * fator

        novo_bid = min(round(novo_bid, 2), bid_maximo)

        if novo_bid == bid_atual:
            continue

        ws.cell(row, bid_c).value = novo_bid
        ws.cell(row, op_c).value = "update"
        relatorio.append({
            "Campanha": camp,
            "Tipo": "Bid",
            "Valor Antigo": bid_atual,
            "Valor Novo": novo_bid,
            "Motivo": f"[{sheet_name}] {motivo}",
        })
        ajustados += 1

    return ajustados


# ---------------------------------------------------------------------------
# Módulo 2 — Budget (uma verba diária por aba)
# ---------------------------------------------------------------------------

def _modulo_budget_sheet(ws_local, cfg_sheet, budget_diario, relatorio, cfg):
    if budget_diario <= 0:
        return 0

    cmap = _resolve_bulk_columns(ws_local, cfg_sheet)
    entity_c = cmap["entity"]
    op_c = cmap["operation"]
    roas_target = cfg["roas_target"]
    budget_minimo = cfg["budget_minimo"]

    campanhas = []
    for row in range(2, ws_local.max_row + 1):
        if ws_local.cell(row, entity_c).value != "Campaign":
            continue
        campanhas.append({
            "sheet_name": cfg_sheet["sheet_name"],
            "ws": ws_local,
            "row": row,
            "camp": _campaign_display_name(ws_local, row, cmap, cfg_sheet["campaign_col"]),
            "budget_col": cfg_sheet["budget_col"],
            "budget": _to_float(ws_local.cell(row, cfg_sheet["budget_col"]).value),
            "sales": _to_float(ws_local.cell(row, cfg_sheet["sales_col"]).value),
            "roas": _to_float(ws_local.cell(row, cfg_sheet["roas_col"]).value),
        })

    if not campanhas:
        return 0

    total_sales = sum(c["sales"] for c in campanhas)
    if total_sales <= 0:
        total_sales = len(campanhas)

    for c in campanhas:
        sales_share = (c["sales"] / total_sales) if total_sales > 0 else (1.0 / len(campanhas))
        fator, motivo = calcular_ajuste_roas(c["roas"], roas_target)

        novo = c["budget"] * fator
        novo = max(novo, budget_minimo)

        budget_max = max(sales_share * budget_diario * 1.5, budget_minimo)
        novo = min(novo, budget_max)

        c["novo_budget"] = round(novo, 2)
        c["motivo"] = (
            f"Aba={c['sheet_name']} | Sales Share={sales_share*100:.1f}% | "
            f"ROAS={c['roas']:.2f} | {motivo}"
        )

    soma = sum(c["novo_budget"] for c in campanhas)
    if soma > budget_diario:
        campanhas_ruins = sorted(
            [c for c in campanhas if c["roas"] < roas_target and c["novo_budget"] > 0],
            key=lambda x: (x["roas"], x["novo_budget"]),
        )
        for c in campanhas_ruins:
            if soma <= budget_diario:
                break
            soma -= c["novo_budget"]
            c["novo_budget"] = 0.0
            c["motivo"] += " | Pausar campanha por baixo ROAS para respeitar budget diário"

    soma = sum(c["novo_budget"] for c in campanhas)
    if soma > budget_diario and soma > 0:
        fator_global = budget_diario / soma
        for c in campanhas:
            if c["novo_budget"] <= 0:
                continue
            c["novo_budget"] = round(c["novo_budget"] * fator_global, 2)
            c["motivo"] += f" | Redução proporcional (fator={fator_global:.4f})"

    soma = round(sum(c["novo_budget"] for c in campanhas), 2)
    sobra = round(budget_diario - soma, 2)
    if sobra > 0:
        ativos = [c for c in campanhas if c["novo_budget"] > 0]
        if ativos:
            pesos = [max(c["roas"], 0.01) for c in ativos]
            peso_total = sum(pesos)
            restante = sobra
            for idx, c in enumerate(ativos):
                if idx == len(ativos) - 1:
                    adicional = restante
                else:
                    adicional = round(sobra * (pesos[idx] / peso_total), 2)
                    restante = round(restante - adicional, 2)
                c["novo_budget"] = round(c["novo_budget"] + adicional, 2)
                c["motivo"] += " | Ajuste para utilização total da verba diária"

    soma_final = round(sum(c["novo_budget"] for c in campanhas), 2)
    diferenca = round(budget_diario - soma_final, 2)
    if diferenca != 0:
        candidatos = [c for c in campanhas if c["novo_budget"] > 0]
        if candidatos:
            melhor = max(candidatos, key=lambda x: x["roas"])
            melhor["novo_budget"] = round(max(melhor["novo_budget"] + diferenca, 0.0), 2)
            melhor["motivo"] += " | Ajuste de centavos para fechar budget diário"

    ajustados = 0
    for c in campanhas:
        novo_budget = round(c["novo_budget"], 2)
        if novo_budget == c["budget"]:
            continue
        c["ws"].cell(c["row"], c["budget_col"]).value = novo_budget
        c["ws"].cell(c["row"], op_c).value = "update"
        relatorio.append({
            "Campanha": c["camp"],
            "Tipo": "Budget",
            "Valor Antigo": c["budget"],
            "Valor Novo": novo_budget,
            "Motivo": c["motivo"],
        })
        ajustados += 1

    return ajustados


# ---------------------------------------------------------------------------
# Módulo 3 — Placement
# ---------------------------------------------------------------------------

def _modulo_placement(ws, relatorio, cfg):
    ajustados   = 0
    roas_target = cfg["roas_target"]

    for row in range(2, ws.max_row + 1):
        entity = ws.cell(row, _COL_ENTITY).value
        if entity != "Bidding Adjustment":
            continue

        placement_type = ws.cell(row, _COL_PLACEMENT_TYPE).value
        if not placement_type:
            # Linha de ajuste de audiência — ignorar
            continue

        pct_atual = _to_float(ws.cell(row, _COL_PLACEMENT_PCT).value)
        roas      = _to_float(ws.cell(row, _COL_ROAS).value)
        label     = f"{_campaign_name(ws, row)} | {placement_type}"

        nova_pct = pct_atual
        motivo   = ""

        if roas > roas_target:
            if pct_atual == 0:
                nova_pct = 10.0
                motivo   = f"ROAS {roas:.2f} acima do target, placement=0 → definido para 10%"
            else:
                fator, motivo = calcular_ajuste_roas(roas, roas_target)
                nova_pct = pct_atual * fator
        elif roas < roas_target:
            if roas == 0 and pct_atual > 0:
                nova_pct = 0.0
                motivo   = "ROAS = 0 → placement zerado"
            elif roas > 0:
                fator, motivo = calcular_ajuste_roas(roas, roas_target)
                nova_pct = pct_atual * fator

        nova_pct = round(max(0.0, min(nova_pct, _PLACEMENT_MAXIMO)), 1)

        if nova_pct == pct_atual:
            continue

        ws.cell(row, _COL_PLACEMENT_PCT).value = nova_pct
        ws.cell(row, _COL_OPERATION).value     = "update"
        relatorio.append({
            "Campanha": label, "Tipo": "Placement",
            "Valor Antigo": pct_atual, "Valor Novo": nova_pct,
            "Motivo": motivo,
        })
        ajustados += 1

    return ajustados


# ---------------------------------------------------------------------------
# Ponto de entrada principal
# ---------------------------------------------------------------------------

def rodar_calibragem(
    arquivo,
    roas_target: float = 4.0,
    budget_diario_sp: float = 500.0,
    budget_diario_sb: float = 0.0,
    budget_diario_sd: float = 0.0,
    bid_maximo: float = 5.0,
    budget_minimo: float = 10.0,
    dias: int = 30,
    calibrar_bid: bool = True,
    calibrar_budget: bool = True,
    calibrar_placement: bool = True,
    on_progress=None,
) -> dict:
    """
    Executa o pipeline completo de calibragem Amazon Ads.

    Args:
        arquivo:               Caminho (str/Path), bytes ou file-like object do .xlsx.
        roas_target:           ROAS alvo das campanhas.
        budget_diario_sp:      Budget diário (teto) só para campanhas em Sponsored Products.
        budget_diario_sb:      Idem para Sponsored Brands (0 = não calibrar budget nesta aba).
        budget_diario_sd:      Idem para Sponsored Display (0 = não calibrar budget nesta aba).
        bid_maximo:            Bid máximo permitido (R$).
        budget_minimo:         Budget mínimo por campanha (R$).
        dias:                  Período de análise para regra de baixo volume (bids).
        calibrar_bid:          Ativar Módulo 1 — Bid (SP, SB e SD quando a aba existir).
        calibrar_budget:       Ativar Módulo 2 — Budget por aba, respeitando cada verba diária.
        calibrar_placement:    Ativar Módulo 3 — Placement (apenas na aba SP).
        on_progress:           Callable(pct: float, msg: str) para atualizar UI (opcional).

    Returns:
        {
          "n_bids":          int,
          "n_budgets":       int,
          "n_placements":    int,
          "relatorio":       list[dict],
          "workbook":        openpyxl.Workbook,
          "abas_removidas":  list[str],
        }

    Raises:
        ValueError: se a aba "Sponsored Products Campaigns" não existir.
    """

    def _prog(pct, msg=""):
        if on_progress:
            on_progress(pct, msg)

    cfg = {
        "roas_target": roas_target,
        "bid_maximo": bid_maximo,
        "budget_minimo": budget_minimo,
        "dias": dias,
    }

    # --- Carregar workbook ---
    _prog(0.05, "Carregando planilha...")
    if isinstance(arquivo, bytes):
        arquivo = io.BytesIO(arquivo)
    elif isinstance(arquivo, (str, Path)):
        pass  # openpyxl aceita diretamente
    wb = openpyxl.load_workbook(arquivo, data_only=True)

    # --- Remover abas RAS ---
    _prog(0.15, "Verificando abas RAS...")
    abas_removidas = []
    for aba in _ABAS_RAS:
        if aba in wb.sheetnames:
            del wb[aba]
            abas_removidas.append(aba)

    # --- Validar aba principal ---
    if _ABA_SP not in wb.sheetnames:
        raise ValueError(
            f"Aba '{_ABA_SP}' não encontrada. "
            f"Abas disponíveis: {wb.sheetnames}"
        )
    ws = wb[_ABA_SP]

    relatorio = []
    n_bids = 0
    n_budgets = 0
    n_placements = 0

    budget_por_chave = {
        "sp": budget_diario_sp,
        "sb": budget_diario_sb,
        "sd": budget_diario_sd,
    }

    # --- Módulo 1: Bid ---
    if calibrar_bid:
        _prog(0.28, "Módulo 1: Calibrando Bids...")
        for key in ("sp", "sb", "sd"):
            sc = _SHEET_CALIBRATION_CONFIG[key]
            name = sc["sheet_name"]
            if name not in wb.sheetnames:
                continue
            n_bids += _modulo_bid_sheet(wb[name], name, relatorio, cfg, sc)

    # --- Módulo 2: Budget (verba independente por aba) ---
    if calibrar_budget:
        _prog(0.55, "Módulo 2: Calibrando Budgets...")
        for key, limite in budget_por_chave.items():
            sc = _SHEET_CALIBRATION_CONFIG[key]
            if sc["sheet_name"] not in wb.sheetnames:
                continue
            n_budgets += _modulo_budget_sheet(wb[sc["sheet_name"]], sc, limite, relatorio, cfg)

    # --- Módulo 3: Placement (somente SP) ---
    if calibrar_placement:
        _prog(0.82, "Módulo 3: Calibrando Placements...")
        n_placements = _modulo_placement(ws, relatorio, cfg)

    _prog(1.0, "Concluído!")

    return {
        "n_bids": n_bids,
        "n_budgets": n_budgets,
        "n_placements": n_placements,
        "relatorio": relatorio,
        "workbook": wb,
        "abas_removidas": abas_removidas,
    }
